
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import calendar
import datetime as dt
from domain.models import CalculationResult

class ExcelExporter:
    def __init__(self, result: CalculationResult, holiday_source: str = ""):
        self.result = result
        self.holiday_source = holiday_source
        self.wb = Workbook()
        self.wb.remove(self.wb.active)

    def export(self, filename: str):
        self._create_calendar_sheet()
        self._create_details_sheet()
        self.wb.save(filename)


    def _create_calendar_sheet(self):
        ws = self.wb.create_sheet('Calendar', 0)
        
        # === CLASSIC STYLES (MATCHING TEMPLATE) ===
        # Fonts
        FONT_TITLE = Font(name='Calibri', size=22, bold=True, color='FFFFFF')
        FONT_HEADER = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        FONT_DAY_NUM = Font(name='Calibri', size=11, bold=True, color='000000') # Day Num same size
        FONT_CONTENT = Font(name='Calibri', size=9) # Inspection says 9.0
        FONT_CONTENT_BOLD = Font(name='Calibri', size=9, bold=True)
        
        # Fills
        FILL_TITLE = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        FILL_HEADER = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        FILL_WEEKEND = PatternFill(start_color='E7F0F7', end_color='E7F0F7', fill_type='solid')
        FILL_PH = PatternFill(start_color='FFF0F0', end_color='FFF0F0', fill_type='solid')
        FILL_PL = PatternFill(start_color='FFF8E7', end_color='FFF8E7', fill_type='solid')
        FILL_WORKING = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        FILL_PADDING = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid') # From inspection A3
        
        # Borders
        BORDER_THIN = Border(
            left=Side(style='thin', color='CCCCCC'), 
            right=Side(style='thin', color='CCCCCC'), 
            top=Side(style='thin', color='CCCCCC'), 
            bottom=Side(style='thin', color='CCCCCC')
        )
        BORDER_BOX = Border(
            left=Side(style='medium', color='7F8C8D'), 
            right=Side(style='medium', color='7F8C8D'), 
            top=Side(style='medium', color='7F8C8D'), 
            bottom=Side(style='medium', color='7F8C8D')
        )
        
        # === LAYOUT SETUP ===
        ws.page_setup.orientation = 'landscape'
        ws.sheet_view.showGridLines = False
        
        # Set column widths (Wide enough for text)
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 22
            
        current_row = 1
        
        # === TITLE BLOCK ===
        ws.merge_cells(f'A{current_row}:G{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = f"{calendar.month_name[self.result.month].upper()} {self.result.year}"
        cell.font = FONT_TITLE
        cell.fill = FILL_TITLE
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[current_row].height = 40
        current_row += 1
        
        # === HEADERS ===
        days = ['MONDAY', 'TUESDAY', 'WEDNESDAY', 'THURSDAY', 'FRIDAY', 'SATURDAY', 'SUNDAY']
        ws.row_dimensions[current_row].height = 25
        for idx, day in enumerate(days, 1):
            c = ws.cell(row=current_row, column=idx, value=day)
            c.font = FONT_HEADER
            c.fill = FILL_HEADER
            c.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # === CALENDAR GRID ===
        cal = calendar.Calendar(firstweekday=0)
        matrix = cal.monthdayscalendar(self.result.year, self.result.month)
        
        holidays_map = {h.date: h.name for h in self.result.holidays}
        leave_map = set(self.result.leave_days)
        working_set = set(self.result.working_days)
        
        for week in matrix:
            ws.row_dimensions[current_row].height = 85 # Boxy square look
            for col_idx, day in enumerate(week):
                cell = ws.cell(row=current_row, column=col_idx + 1)
                cell.border = BORDER_THIN
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                
                if day == 0:
                    cell.fill = FILL_PADDING
                    continue
                    
                date_obj = dt.date(self.result.year, self.result.month, day)
                
                # Determine Content & Style
                bg_fill = FILL_WEEKEND if col_idx >= 5 else FILL_WORKING
                content_text = ""
                content_font = FONT_CONTENT
                
                if date_obj in leave_map:
                    bg_fill = FILL_PL
                    content_text = "PL"
                    content_font = FONT_CONTENT_BOLD
                elif date_obj in holidays_map:
                    bg_fill = FILL_PH
                    content_text = f"PH\n{holidays_map[date_obj]}"
                    content_font = FONT_CONTENT_BOLD
                elif col_idx >= 5: # Sat/Sun
                    bg_fill = FILL_WEEKEND
                    content_text = "WE"
                    content_font = FONT_CONTENT
                elif date_obj in working_set:
                    content_text = "WD"
                    content_font = FONT_CONTENT
                    
                # Apply Fill
                cell.fill = bg_fill
                
                # Match Template Loop: Day Num (Bold 11) then Content (Regular 9)
                # openpyxl limitation: setting cell.font affects whole cell.
                # Template inspection showed D3 (PH) font as Calibri 9.0.
                # A2 (Mon) was Calibri 11.
                # So we should probably default the CELL font to the content font (Size 9).
                # The Day number being 11 Bold in the inspection might have been applied via RichText or separate inspection?
                # Wait, inspection for D3 said "Font: Calibri 9.0". D3 contained "1\nPH\n..."
                # Use font size 9 for the whole cell to match D3.
                
                cell.value = f"{day}\n{content_text}"
                cell.font = content_font
                
                # If we could, we would use RichText for bolding just the number, 
                # but openpyxl's rich text support is tricky. 
                # Simple approach: The whole cell takes the content font.
                # However, we want the number to be big.
                # Let's stick to consistent font for now to match the user's "clean" request.
                
            current_row += 1
            
        # === SUMMARY FOOTER (LEGACY) ===
        current_row += 1
        
        # 1. SUMMARY Header Bar
        ws.merge_cells(f'A{current_row}:G{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = "SUMMARY"
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[current_row].height = 20
        current_row += 1
        
        # 2. Data Row (Eligible Days | Exclusions | Grade | Total)
        ws.row_dimensions[current_row].height = 25
        
        # A-B: Eligible Days
        ws.merge_cells(f'A{current_row}:B{current_row}')
        c_elig = ws[f'A{current_row}']
        # Rich text would be ideal, but for simplicity we'll just format the cell text
        c_elig.value = f"Eligible Working Days: {self.result.working_days_count}"
        c_elig.font = Font(name='Calibri', size=11, bold=True)
        c_elig.alignment = Alignment(horizontal='left', vertical='center')
        
        # C-D: Exclusions
        ws.merge_cells(f'C{current_row}:D{current_row}')
        c_excl = ws[f'C{current_row}']
        excl_text = f"Exclusions: WE={self.result.weekend_days_count} PH={self.result.public_holidays_count} PL={self.result.personal_leave_count}"
        c_excl.value = excl_text
        c_excl.font = Font(name='Calibri', size=10, color='666666')
        c_excl.alignment = Alignment(horizontal='center', vertical='center')
        
        # E-F: Grade
        ws.merge_cells(f'E{current_row}:F{current_row}')
        c_grade = ws[f'E{current_row}']
        c_grade.value = f"Grade: {self.result.grade} @ RM{self.result.rate}/day"
        c_grade.font = Font(name='Calibri', size=10, bold=True)
        c_grade.alignment = Alignment(horizontal='right', vertical='center')
        
        # G: Total
        c_total = ws[f'G{current_row}']
        c_total.value = f"TOTAL: RM {self.result.total_reimbursement:,.2f}"
        c_total.font = Font(name='Calibri', size=12, bold=True, color='C00000') # Red
        c_total.alignment = Alignment(horizontal='right', vertical='center')
        
        current_row += 2
        
        # 3. Legend & Info
        # Legend
        ws.merge_cells(f'A{current_row}:G{current_row}')
        c_leg = ws[f'A{current_row}']
        c_leg.value = "Legend: WD Working Day | WE Weekend | PH Public Holiday | PL Personal Leave"
        c_leg.font = Font(name='Calibri', size=9, color='666666')
        current_row += 1
        
        # All Grades Ref
        ws.merge_cells(f'A{current_row}:G{current_row}')
        c_ref = ws[f'A{current_row}']
        # Reconstruct grade str
        from domain.calculator import GRADE_RATES
        rates_str = " ".join([f"{k} RM{v:,}" for k,v in GRADE_RATES.items()])
        c_ref.value = f"All Grades (This Month): {rates_str}"
        c_ref.font = Font(name='Calibri', size=9, color='888888')
        current_row += 1
        
        # Source
        ws.merge_cells(f'A{current_row}:G{current_row}')
        c_src = ws[f'A{current_row}']
        c_src.value = f"Holiday source: {self.holiday_source}"
        c_src.font = Font(name='Calibri', size=8, italic=True, color='AAAAAA')

    def _create_details_sheet(self):
        ws = self.wb.create_sheet('Details', 1)
        ws.append(["Date", "Day", "Status", "Note"])
        
        # We need to iterate all days
        cal = calendar.Calendar()
        for day in cal.itermonthdates(self.result.year, self.result.month):
            if day.month != self.result.month:
                continue
                
            status = "Working Day"
            note = ""
            
            holidays_map = {h.date: h.name for h in self.result.holidays}
            
            if day in holidays_map:
                status = "Public Holiday"
                note = holidays_map[day]
            elif day in self.result.leave_days:
                status = "Personal Leave"
            elif day.weekday() >= 5:
                status = "Weekend"
            elif day not in self.result.working_days:
                 status = "Excluded"
            
            ws.append([day, day.strftime("%A"), status, note])
